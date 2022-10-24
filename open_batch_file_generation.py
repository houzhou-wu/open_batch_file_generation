# %%
import sys
import codecs
from pathlib import Path
from shutil import copy, copytree
from datetime import datetime
from pydocxtpl import DocxWriter
from xltpl.writerx import BookWriter
from openpyxl import load_workbook
from pandas import read_excel, DataFrame
from gooey import Gooey, GooeyParser
# %%
def io_encode():
    if sys.stdout.encoding != 'UTF-8':
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if sys.stderr.encoding != 'UTF-8':
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')
# %%
def print_dash(n):
    print('-'*n, flush=True)
# %%
def print_message(args):
    print(datetime.now(), args.Data, flush=True)
    print(datetime.now(), args.Template, flush=True)
    print(datetime.now(), args.Directory, flush=True)
    print(datetime.now(), args.multi_row_file, flush=True)
# %%
def render_docx(args, main_dict, sub_records, file):
    if args.multi_row_file != None:
        main_dict['records'] = sub_records
    tplt = DocxWriter(file)
    tplt.render(main_dict)
    tplt.save(file)
# %%
def render_xlsx(args, main_dict, sub_records, file):
    if args.multi_row_file != None:
        main_dict['records'] = sub_records
    tplt = BookWriter(file)
    for sheet_name in load_workbook(file).sheetnames:
        main_dict['tpl_name'] = sheet_name
        main_dict['sheet_name'] = sheet_name
        tplt.render_sheet(main_dict)
    tplt.save(file)
# %%
def process4single_file(args, main_records, sub_records, template_file, save_directory):
    for main_dict in main_records:
        file = save_directory / \
            (list(main_dict.values())[0] + str(template_file)[-5:])
        copy(template_file, file)
        if args.multi_row_file != None:
            df = DataFrame.from_records(sub_records)
            df = df[df[list(main_dict.keys())[0]] ==
                    list(main_dict.values())[0]]
            subs_records = df.to_records()
        else:
            subs_records = None
        if template_file.suffix == '.docx':
            render_docx(args, main_dict, subs_records, file)
        elif template_file.suffix == '.xlsx':
            render_xlsx(args, main_dict, subs_records, file)
        print(datetime.now(), file, flush=True)
# %%
def process4one_directory(args, main_records, sub_records, template_directory, save_directory):
    for main_dict in main_records:
        dir = save_directory / list(main_dict.values())[0]
        copytree(template_directory, dir)
        for file in dir.rglob('*'):
            if args.multi_row_file != None:
                df = DataFrame.from_records(sub_records)
                sf = df[df[list(main_dict.keys())[0]] ==
                        list(main_dict.values())[0]]
                subs_records = sf.to_records()
            else:
                subs_records = None
            if file.suffix == '.docx':
                render_docx(args, main_dict, subs_records, file)
            elif file.suffix == '.xlsx':
                render_xlsx(args, main_dict, subs_records, file)
            print(datetime.now(), file, flush=True)
# %%
def process4main(args):
    main_records = read_excel(args.Data,
                              sheet_name=load_workbook(args.Data).active.title).to_dict(orient='records')
    if args.multi_row_file != None:
        sub_records = read_excel(args.multi_row_file,
                                 sheet_name=load_workbook(args.multi_row_file).active.title).to_dict(orient='records')
    else:
        sub_records = None
    if args.command == 'SingleFile':
        print('Actions: SingleFile', flush=True)
        print_dash(64)
        print_message(args)
        print_dash(64)
        process4single_file(args, main_records, sub_records,
                            Path(args.Template), Path(args.Directory))
        print_dash(64)
    elif args.command == 'OneDirectory':
        print('Actions: OneDirectory', flush=True)
        print_dash(64)
        print_message(args)
        print_dash(64)
        process4one_directory(args, main_records, sub_records, 
                            Path(args.Template), Path(args.Directory))
        print_dash(64)
# %%
@ Gooey(optional_cols=2, program_name='Open Batch File Generation')
def main():
    parser = GooeyParser(
        description='Open batch file generation for Single File, One Directory')
    subs = parser.add_subparsers(dest='command')

    single_file_parser = subs.add_parser(
        'SingleFile')
    single_file_parser.add_argument('Data',
                                    metavar='Data File',
                                    help='File to select',
                                    widget='FileChooser',
                                    type=str)
    single_file_parser.add_argument('Template',
                                    metavar='Template File',
                                    help='File to select',
                                    widget='FileChooser',
                                    type=str)
    single_file_parser.add_argument('Directory',
                                    metavar='Save Directory',
                                    help='Directory to select',
                                    widget='DirChooser',
                                    type=str)
    single_file_parser.add_argument('--multi_row_file',
                                    metavar='multi row file',
                                    help='File to select',
                                    widget='FileChooser',
                                    type=str)

    one_directory_parser = subs.add_parser(
        'OneDirectory')
    one_directory_parser.add_argument('Data',
                                      metavar='Data File',
                                      help='File to select',
                                      widget='FileChooser',
                                      type=str)
    one_directory_parser.add_argument('Template',
                                      metavar='Template Directory',
                                      help='Directory to select',
                                      widget='DirChooser',
                                      type=str)
    one_directory_parser.add_argument('Directory',
                                      metavar='Save Directory',
                                      help='Directory to select',
                                      widget='DirChooser',
                                      type=str)
    one_directory_parser.add_argument('--multi_row_file',
                                      metavar='multi row file',
                                      help='File to select',
                                      widget='FileChooser',
                                      type=str)

    process4main(parser.parse_args())
# %%
if __name__ == '__main__':
    io_encode()
    main()